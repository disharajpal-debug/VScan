from flask import (
    Flask,
    request,
    jsonify,
    render_template,
    redirect,
    url_for,
    session,
    flash,
    make_response,
)
from flask_cors import CORS
import os
import base64
import requests
import json
import datetime
from pymongo import MongoClient
from werkzeug.utils import secure_filename
from bson import ObjectId
from functools import wraps
from flask_bcrypt import Bcrypt
import re
import secrets
import smtplib
import time
import csv
import io
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dotenv import load_dotenv
from google_sheets_integration import append_to_master_sheet

# Load environment variables
load_dotenv()
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors

app = Flask(__name__)
# Secure Flask secret key from environment
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev_secret_key_change_me")
CORS(app)
bcrypt = Bcrypt(app)

# Configuration
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png"}

# API Configuration - Load from environment
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-2.0-flash")
GEMINI_API_URL = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"

if not GEMINI_API_KEY:
    raise ValueError(
        "GEMINI_API_KEY not found in environment variables. Please check .env file."
    )


# MongoDB configuration - Load from environment
# MongoDB configuration - MUST come from environment on Render
MONGO_URI = os.environ.get("MONGO_URI")
DB_NAME = os.environ.get("DATABASE_NAME", "visiting_card")
COLLECTION_NAME = os.environ.get("DB_COLLECTION_NAME", "cards")

if not MONGO_URI:
    raise Exception(
        "❌ MONGO_URI is not set! Add it in Render Dashboard → Environment."
    )


# Create upload folder if it doesn't exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# MongoDB client (local, no certifi needed)
client = MongoClient(MONGO_URI)
db = client[DB_NAME]
collection = db[COLLECTION_NAME]
users_collection = db["users"]

reset_tokens = {}

# SMTP configuration - Load from environment
SMTP_SERVER = os.environ.get("SMTP_SERVER", "smtp.example.com")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USERNAME = os.environ.get("SMTP_USERNAME", "your_email@example.com")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "your_email_password")
SENDER_EMAIL = os.environ.get("SENDER_EMAIL", "your_email@example.com")

# Privileged user for special features
PRIVILEGED_USER = os.environ.get("PRIVILEGED_USER", "disha.rajpal@c4i4.com")

# Static admin credentials (unchangeable - set in .env)
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "admin@c4i4.org").lower()
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "AdminPass123!")

# Roles
ROLE_ADMIN = "admin"
ROLE_USER = "user"

# Secure session cookies (adapt for development vs production)
app.config["SESSION_COOKIE_SECURE"] = (
    os.environ.get("SESSION_COOKIE_SECURE", "False").lower() == "true"
)
app.config["SESSION_COOKIE_HTTPONLY"] = (
    os.environ.get("SESSION_COOKIE_HTTPONLY", "True").lower() == "true"
)
app.config["SESSION_COOKIE_SAMESITE"] = os.environ.get("SESSION_COOKIE_SAMESITE", "Lax")


def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def validate_form_fields(fields):
    """Validate that all required fields are present and not empty"""
    for field in fields:
        if not field or (isinstance(field, str) and not field.strip()):
            return False
    return True


def encode_image_to_base64(image_path):
    """Encode image to base64"""
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode("utf-8")


def extract_details_with_gemini(image_base64):
    """Extract details from image using Gemini API with retry logic for rate limiting"""
    headers = {
        "Content-Type": "application/json",
    }

    prompt = """
    Please analyze this visiting card image and extract the following information in JSON format:
    {
        "name": "Full name of the person",
        "company": "Company name",
        "designation": "Job title/position",
        "email": "Email address",
        "phone": "Phone number",
        "address": "Address",
        "website": "Website URL",
        "additional_info": "Any other relevant information"
    }
    
    If any field is not found, set it to null. Return only the JSON object, no additional text.
    """

    data = {
        "contents": [
            {
                "parts": [
                    {"text": prompt},
                    {"inline_data": {"mime_type": "image/jpeg", "data": image_base64}},
                ]
            }
        ]
    }

    # Retry logic with exponential backoff
    max_retries = 3
    retry_delay = 2  # Start with 2 seconds

    for attempt in range(max_retries):
        try:
            response = requests.post(
                f"{GEMINI_API_URL}?key={GEMINI_API_KEY}",
                headers=headers,
                json=data,
                timeout=30,
            )

            if response.status_code == 200:
                result = response.json()
                if "candidates" in result and len(result["candidates"]) > 0:
                    content = result["candidates"][0]["content"]["parts"][0]["text"]
                    # Try to extract JSON from the response
                    try:
                        # Find JSON in the response
                        start_idx = content.find("{")
                        end_idx = content.rfind("}") + 1
                        if start_idx != -1 and end_idx != 0:
                            json_str = content[start_idx:end_idx]
                            return json.loads(json_str)
                        else:
                            return {"error": "No JSON found in response"}
                    except json.JSONDecodeError:
                        return {
                            "error": "Invalid JSON response",
                            "raw_response": content,
                        }
                else:
                    return {"error": "No content in response"}

            # Handle rate limiting (429) and service unavailable (503)
            elif response.status_code == 429:
                if attempt < max_retries - 1:
                    wait_time = retry_delay * (
                        2**attempt
                    )  # Exponential backoff: 2, 4, 8 seconds
                    print(
                        f"Rate limited (429). Retrying in {wait_time} seconds... (Attempt {attempt + 1}/{max_retries})"
                    )
                    time.sleep(wait_time)
                    continue
                else:
                    return {
                        "error": "API rate limit exceeded",
                        "details": "Too many requests. Please try again in a few minutes.",
                        "status": 429,
                    }

            elif response.status_code == 503:
                if attempt < max_retries - 1:
                    wait_time = retry_delay * (2**attempt)
                    print(
                        f"Service unavailable (503). Retrying in {wait_time} seconds... (Attempt {attempt + 1}/{max_retries})"
                    )
                    time.sleep(wait_time)
                    continue
                else:
                    return {
                        "error": "API service temporarily unavailable",
                        "details": "The Gemini API is temporarily unavailable. Please try again later.",
                        "status": 503,
                    }

            else:
                print(
                    f"Gemini API request failed: {response.status_code} {response.text}"
                )
                return {
                    "error": f"API request failed with status {response.status_code}",
                    "details": response.text,
                    "status": response.status_code,
                }

        except requests.exceptions.Timeout:
            if attempt < max_retries - 1:
                wait_time = retry_delay * (2**attempt)
                print(
                    f"Request timeout. Retrying in {wait_time} seconds... (Attempt {attempt + 1}/{max_retries})"
                )
                time.sleep(wait_time)
                continue
            else:
                return {
                    "error": "Request timeout",
                    "details": "The API request took too long. Please try again with a smaller image.",
                }

        except requests.exceptions.RequestException as e:
            if attempt < max_retries - 1:
                wait_time = retry_delay * (2**attempt)
                print(
                    f"Network error: {e}. Retrying in {wait_time} seconds... (Attempt {attempt + 1}/{max_retries})"
                )
                time.sleep(wait_time)
                continue
            else:
                return {
                    "error": "Network error",
                    "details": f"Failed to connect to API: {str(e)}",
                }

    # If all retries failed
    return {
        "error": "Failed after multiple retries",
        "details": "Could not complete the request after 3 attempts. Please try again later.",
    }


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get("role") != ROLE_ADMIN:
            return jsonify({"error": "Admin access required"}), 403
        return f(*args, **kwargs)

    return decorated_function


def get_user_by_email(email):
    return users_collection.find_one({"email": email})


def is_strong_password(password):
    # At least 8 chars, one upper, one lower, one digit, one special char
    if (
        len(password) < 8
        or not re.search(r"[A-Z]", password)
        or not re.search(r"[a-z]", password)
        or not re.search(r"\d", password)
        or not re.search(r"[^A-Za-z0-9]", password)
    ):
        return False
    return True


def get_user_by_employee_id(employee_id):
    return users_collection.find_one({"employee_id": employee_id})


@app.route("/login", methods=["GET", "POST"])
def login():
    # Redirect to index if already logged in
    if "user" in session:
        return redirect(url_for("index"))

    if request.method == "POST":
        login_input = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        # Validate form fields
        if not validate_form_fields([login_input, password]):
            flash("Email/Employee ID and password are required.", "danger")
            return render_template("login.html")

        # Check for static admin login first
        if login_input == ADMIN_EMAIL and password == ADMIN_PASSWORD:
            session["user"] = ADMIN_EMAIL
            session["role"] = ROLE_ADMIN
            session["name"] = "Administrator"
            session["show_welcome_modal"] = True
            session.permanent = False
            return redirect(url_for("admin_dashboard"))

        # Regular user login (database)
        user = None
        if "@" in login_input:
            user = get_user_by_email(login_input)
        else:
            user = get_user_by_employee_id(login_input)

        if user and bcrypt.check_password_hash(user["password"], password):
            session["user"] = user["email"]
            session["role"] = user.get("role", ROLE_USER)
            session["name"] = user.get("name", "User")
            return redirect(url_for("index"))
        else:
            flash("Invalid email/employee ID or password", "danger")
    return render_template("login.html")


@app.route("/logout")
def logout():
    # Clear all session data
    session.clear()
    return render_template("logout_redirect.html")


@app.route("/")
@login_required
def index():
    name = session.get("name", "")
    if name:
        name = name.capitalize()
    show_welcome_modal = session.pop("show_welcome_modal", False)
    if "just_logged_in" in session:
        flash(f"Welcome, {name}!", "success")
        session.pop("just_logged_in")
    return render_template(
        "index.html", name=name, show_welcome_modal=show_welcome_modal
    )


@app.route("/upload", methods=["POST"])
def upload_file():
    try:
        if "file" not in request.files:
            print("No file part in request.files")
            return jsonify({"error": "No file part"}), 400

        file = request.files["file"]
        if file.filename == "":
            print("No selected file")
            return jsonify({"error": "No selected file"}), 400

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            file.save(filepath)

            # Encode image to base64
            image_base64 = encode_image_to_base64(filepath)

            # Extract details using Gemini API
            extracted_data = extract_details_with_gemini(image_base64)

            if "error" in extracted_data:
                print(f"Gemini extraction error: {extracted_data}")
                os.remove(filepath)

                # Return appropriate status code based on error type
                status_code = 400
                if extracted_data.get("status") == 429:
                    status_code = 429  # Too Many Requests
                elif extracted_data.get("status") == 503:
                    status_code = 503  # Service Unavailable

                return (
                    jsonify(
                        {
                            "error": extracted_data["error"],
                            "details": extracted_data.get("details", ""),
                        }
                    ),
                    status_code,
                )

            # Add timestamp, filename, and scanned_by to the data
            extracted_data["uploaded_at"] = datetime.datetime.now()
            extracted_data["original_filename"] = filename
            extracted_data["scanned_by"] = session["user"]
            extracted_data["shared"] = False

            # Stricter duplicate check (by normalized email, phone, or name+company)
            email = extracted_data.get("email", "").strip().lower()
            phone = extracted_data.get("phone", "").strip()
            name = extracted_data.get("name", "").strip().lower()
            company = extracted_data.get("company", "").strip().lower()

            duplicate_query = {"scanned_by": session["user"]}
            or_conditions = []

            if email:
                or_conditions.append({"email": email})
            if phone:
                or_conditions.append({"phone": phone})
            if name and company:
                or_conditions.append({"name": name, "company": company})

            if or_conditions:
                duplicate_query["$or"] = or_conditions
                duplicate = collection.find_one(duplicate_query)
                if duplicate:
                    os.remove(filepath)
                    return (
                        jsonify(
                            {
                                "success": False,
                                "error": "Duplicate card detected. This card already exists.",
                            }
                        ),
                        409,
                    )

            # Check for extract_only mode
            extract_only = request.args.get("extract_only") == "1"
            if extract_only:
                os.remove(filepath)
                return jsonify(
                    {
                        "success": True,
                        "data": extracted_data,
                        "message": "Card details extracted. Confirm to save.",
                    }
                )

            # Save to MongoDB
            result = collection.insert_one(extracted_data)
            extracted_data["_id"] = str(result.inserted_id)

            # Only export to Google Sheets for privileged user
            user_email = session.get("user")
            print(f"Current user: {user_email}, Privileged user: {PRIVILEGED_USER}")
            if user_email == PRIVILEGED_USER:
                try:
                    append_to_master_sheet(extracted_data)
                    print("Successfully exported to Google Sheets (master)")
                except Exception as e:
                    print(f"Failed to export to Google Sheets (master): {e}")

            # Clean up uploaded file
            os.remove(filepath)

            return jsonify(
                {
                    "success": True,
                    "data": extracted_data,
                    "message": "Card details extracted and saved successfully!",
                }
            )

        else:
            print("Invalid file type uploaded")
            return (
                jsonify(
                    {
                        "error": "Invalid file type. Please upload JPG, JPEG, or PNG files."
                    }
                ),
                400,
            )

    except Exception as e:
        print(f"Server error: {e}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500


@app.route("/cards", methods=["GET"])
@login_required
def get_cards():
    """Get cards visible to the current user"""
    try:
        # Allow admin to filter by who scanned the card via query param `scanned_by`
        scanned_by = request.args.get("scanned_by")

        # Admin: can see all cards, or filtered by scanned_by
        if session.get("role") == ROLE_ADMIN:
            query = {}
            if scanned_by:
                query["scanned_by"] = scanned_by
            cards = list(collection.find(query))
        else:
            # Regular users: show cards shared by privileged user or explicitly shared
            query = {"$or": [{"scanned_by": PRIVILEGED_USER}, {"shared": True}]}
            cards = list(collection.find(query))
            # Extra safeguard: filter in Python as well
            cards = [
                card
                for card in cards
                if card.get("scanned_by") == PRIVILEGED_USER
                or card.get("shared") is True
            ]

        # Normalize results for JSON (ObjectId -> str, datetime -> isoformat)
        for card in cards:
            card["_id"] = str(card["_id"])
            if card.get("uploaded_at") and hasattr(
                card.get("uploaded_at"), "isoformat"
            ):
                try:
                    card["uploaded_at"] = card["uploaded_at"].isoformat()
                except Exception:
                    pass
        return jsonify({"success": True, "cards": cards})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch cards: {str(e)}"}), 500


@app.route("/admin/scanners")
@login_required
@admin_required
def admin_scanners():
    """Return distinct scanner identities (email + name if available) for admin filter dropdown"""
    try:
        scanners = collection.distinct("scanned_by")
        result = []
        for s in scanners:
            if not s:
                continue
            user = users_collection.find_one({"email": s})
            name = None
            if user:
                name = user.get("name")
            result.append({"email": s, "name": name or s})
        return jsonify({"success": True, "scanners": result})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch scanners: {str(e)}"}), 500


@app.route("/admin/scanner")
@login_required
@admin_required
def admin_scanner_redirect():
    """Compatibility redirect: singular -> plural endpoint"""
    return redirect(url_for("admin_scanners"))


@app.route("/admin/users")
@login_required
@admin_required
def admin_users():
    """Return all registered users with card count (admin only)"""
    try:
        users = list(users_collection.find({}, {"password": 0}))
        # Convert ObjectId to string and datetime to isoformat, add card count
        for user in users:
            user["_id"] = str(user.get("_id", ""))
            if user.get("created_at") and hasattr(user.get("created_at"), "isoformat"):
                try:
                    user["created_at"] = user["created_at"].isoformat()
                except Exception:
                    pass
            # Count cards scanned by this user (match by user name stored in scanned_by field)
            user_name = user.get("name", "")
            cards_count = (
                collection.count_documents({"scanned_by": user_name})
                if user_name
                else 0
            )
            user["cards_scanned"] = cards_count
        return jsonify({"success": True, "users": users})
    except Exception as e:
        return jsonify({"error": f"Failed to fetch users: {str(e)}"}), 500


@app.route("/update_card/<card_id>", methods=["PUT"])
def update_card(card_id):
    """Update card details in database"""
    try:
        # Validate ObjectId
        if not ObjectId.is_valid(card_id):
            return jsonify({"error": "Invalid card ID"}), 400

        # Get update data from request
        update_data = request.get_json()

        # Remove any fields that shouldn't be updated
        fields_to_remove = ["_id", "uploaded_at", "original_filename"]
        for field in fields_to_remove:
            update_data.pop(field, None)

        # Update the card in MongoDB
        result = collection.update_one(
            {"_id": ObjectId(card_id)}, {"$set": update_data}
        )

        if result.matched_count == 0:
            return jsonify({"error": "Card not found"}), 404

        if result.modified_count == 0:
            return jsonify({"error": "No changes made"}), 400

        return jsonify({"success": True, "message": "Card updated successfully"})

    except Exception as e:
        print(f"Update card error: {e}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500


@app.route("/delete_card/<card_id>", methods=["DELETE"])
def delete_card(card_id):
    """Delete card from database"""
    try:
        # Validate ObjectId
        if not ObjectId.is_valid(card_id):
            return jsonify({"error": "Invalid card ID"}), 400

        # Delete the card from MongoDB
        result = collection.delete_one({"_id": ObjectId(card_id)})

        if result.deleted_count == 0:
            return jsonify({"error": "Card not found"}), 404

        return jsonify({"success": True, "message": "Card deleted successfully"})

    except Exception as e:
        print(f"Delete card error: {e}")
        return jsonify({"error": f"Server error: {str(e)}"}), 500


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        employee_id = request.form.get("employee_id", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        # Validate all fields
        if not all([name, employee_id, email, password, confirm_password]):
            flash("All fields are required.", "danger")
            return render_template(
                "signup.html", name=name, employee_id=employee_id, email=email
            )

        # Passwords match
        if password != confirm_password:
            flash("Passwords do not match.", "danger")
            return render_template(
                "signup.html", name=name, employee_id=employee_id, email=email
            )

        # Password strength
        if not is_strong_password(password):
            flash(
                "Password must be at least 8 characters long and include uppercase, lowercase, digit, and special character.",
                "danger",
            )
            return render_template(
                "signup.html", name=name, employee_id=employee_id, email=email
            )

        # Unique email
        if get_user_by_email(email):
            flash("Email already registered.", "danger")
            return render_template(
                "signup.html", name=name, employee_id=employee_id, email=email
            )

        # Unique employee ID
        if get_user_by_employee_id(employee_id):
            flash("Employee ID already registered.", "danger")
            return render_template(
                "signup.html", name=name, employee_id=employee_id, email=email
            )

        # Hash password and insert
        hashed_password = bcrypt.generate_password_hash(password).decode("utf-8")
        user = {
            "name": name,
            "employee_id": employee_id,
            "email": email,
            "password": hashed_password,
            "created_at": datetime.datetime.utcnow(),
            "role": "user",
        }
        users_collection.insert_one(user)
        flash("Registration successful! Please log in.", "success")
        return redirect(url_for("login"))

    return render_template("signup.html")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    # New flow: user submits their email, then is immediately shown a form
    # to choose a new password (no external email is sent).
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()

        # Prevent admin from resetting password through this endpoint
        if email == ADMIN_EMAIL:
            flash(
                "Admin password cannot be reset through this form. Please contact system administrator.",
                "danger",
            )
            return render_template("forgot_password.html")

        user = get_user_by_email(email)
        if not user:
            flash("No account found with that email.", "danger")
            return render_template("forgot_password.html")

        # Render the reset-password form where user will enter new password
        return render_template("reset_password.html", email=email)

    return render_template("forgot_password.html")


@app.route("/reset-password", methods=["POST"])
def reset_password():
    # Handle password update after user submits new password
    email = request.form.get("email", "").strip().lower()
    new_password = request.form.get("new_password", "")
    confirm_password = request.form.get("confirm_password", "")

    # Basic validations
    if not email or not new_password or not confirm_password:
        flash("All fields are required.", "danger")
        return render_template("reset_password.html", email=email)

    if new_password != confirm_password:
        flash("New passwords do not match.", "danger")
        return render_template("reset_password.html", email=email)

    if not is_strong_password(new_password):
        flash(
            "Password must be at least 8 characters long and include uppercase, lowercase, digit, and special character.",
            "danger",
        )
        return render_template("reset_password.html", email=email)

    # Prevent changing admin via this flow
    if email == ADMIN_EMAIL:
        flash(
            "Admin password cannot be reset through this form. Please contact system administrator.",
            "danger",
        )
        return redirect(url_for("forgot_password"))

    user = get_user_by_email(email)
    if not user:
        flash("No account found with that email.", "danger")
        return redirect(url_for("forgot_password"))

    # Hash and update
    hashed_password = bcrypt.generate_password_hash(new_password).decode("utf-8")
    users_collection.update_one(
        {"email": email}, {"$set": {"password": hashed_password}}
    )

    flash("Password changed successfully! Please log in.", "success")
    return redirect(url_for("login"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    """Allow logged-in users to change their password"""
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        confirm_password = request.form.get("confirm_password", "")

        user_email = session.get("user")
        user = get_user_by_email(user_email)

        if not user:
            flash("User not found.", "danger")
            return render_template("change_password.html")

        # Verify current password
        if not bcrypt.check_password_hash(user.get("password", ""), current_password):
            flash("Current password is incorrect.", "danger")
            return render_template("change_password.html")

        if not new_password or not confirm_password:
            flash("All fields are required.", "danger")
            return render_template("change_password.html")

        if new_password != confirm_password:
            flash("New passwords do not match.", "danger")
            return render_template("change_password.html")

        if not is_strong_password(new_password):
            flash(
                "Password must be at least 8 characters long and include uppercase, lowercase, digit, and special character.",
                "danger",
            )
            return render_template("change_password.html")

        # Update password in database
        hashed_password = bcrypt.generate_password_hash(new_password).decode("utf-8")
        users_collection.update_one(
            {"email": user_email}, {"$set": {"password": hashed_password}}
        )

        flash("Password changed successfully!", "success")
        return redirect(url_for("login"))

    return render_template("change_password.html")


@app.route("/personal-cards")
@login_required
def personal_cards():
    user_email = session.get("user")
    # Backend safeguard: Only show cards if scanned_by matches the current user exactly
    # Privileged user doesn't see personal cards (they see shared cards instead)
    if not user_email or user_email == PRIVILEGED_USER:
        cards = []
    else:
        cards = list(collection.find({"scanned_by": user_email}))
        # Extra safeguard: filter in Python as well
        cards = [card for card in cards if card.get("scanned_by") == user_email]
        for card in cards:
            card["_id"] = str(card["_id"])
    if request.args.get("json") == "1":
        return jsonify({"success": True, "cards": cards})
    return render_template("personal_cards.html", cards=cards)


@app.route("/share_card/<card_id>", methods=["POST"])
@login_required
def share_card(card_id):
    """Toggle the shared status of a personal card"""
    try:
        if not ObjectId.is_valid(card_id):
            return jsonify({"error": "Invalid card ID"}), 400
        card = collection.find_one({"_id": ObjectId(card_id)})
        if not card:
            return jsonify({"error": "Card not found"}), 404
        user_email = session.get("user")
        user_role = session.get("role", ROLE_USER)

        # Allow admins to share any card, or users to share their own cards
        if user_role != ROLE_ADMIN and card.get("scanned_by") != user_email:
            return jsonify({"error": "Unauthorized"}), 403

        new_shared = not card.get("shared", False)
        collection.update_one(
            {"_id": ObjectId(card_id)}, {"$set": {"shared": new_shared}}
        )
        return jsonify({"success": True, "shared": new_shared})
    except Exception as e:
        return jsonify({"error": f"Server error: {str(e)}"}), 500


@app.route("/admin/export_csv")
@login_required
@admin_required
def admin_export_csv():
    """Export all cards as CSV (admin only)"""
    try:
        # Allow exporting filtered by scanner when `scanned_by` query param is present
        scanned_by = request.args.get("scanned_by")
        query = {}
        if scanned_by:
            query["scanned_by"] = scanned_by
        cards = list(collection.find(query))

        # Prepare CSV
        output = io.StringIO()
        writer = csv.writer(output)
        headers = [
            "Name",
            "Company",
            "Designation",
            "Email",
            "Phone",
            "Website",
            "Address",
            "Additional Info",
            "Scanned By",
            "Shared",
            "Uploaded At",
        ]
        writer.writerow(headers)

        for c in cards:
            uploaded_str = ""
            if c.get("uploaded_at"):
                try:
                    uploaded_str = c.get("uploaded_at").isoformat()
                except Exception:
                    uploaded_str = str(c.get("uploaded_at"))
            row = [
                c.get("name", ""),
                c.get("company", ""),
                c.get("designation", ""),
                c.get("email", ""),
                c.get("phone", ""),
                c.get("website", ""),
                c.get("address", ""),
                c.get("additional_info", ""),
                c.get("scanned_by", ""),
                "Yes" if c.get("shared") else "No",
                uploaded_str,
            ]
            writer.writerow(row)

        output.seek(0)
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = "attachment; filename=all_cards.csv"
        response.headers["Content-Type"] = "text/csv"
        return response
    except Exception as e:
        print(f"Export CSV error: {e}")
        return jsonify({"error": f"Failed to export CSV: {str(e)}"}), 500


@app.route("/admin/export_excel")
@login_required
@admin_required
def admin_export_excel():
    """Export all cards as Excel (admin only)"""
    try:
        cards = list(collection.find({}))

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cards"

        # Define headers
        headers = [
            "Name",
            "Company",
            "Designation",
            "Email",
            "Phone",
            "Website",
            "Address",
            "Additional Info",
            "Scanned By",
            "Shared",
            "Uploaded At",
        ]

        # Add header row with formatting
        header_fill = PatternFill(
            start_color="667eea", end_color="667eea", fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data rows
        for row_num, card in enumerate(cards, 2):
            uploaded_str = ""
            if card.get("uploaded_at"):
                try:
                    uploaded_str = card.get("uploaded_at").isoformat()
                except Exception:
                    uploaded_str = str(card.get("uploaded_at"))

            row_data = [
                card.get("name", ""),
                card.get("company", ""),
                card.get("designation", ""),
                card.get("email", ""),
                card.get("phone", ""),
                card.get("website", ""),
                card.get("address", ""),
                card.get("additional_info", ""),
                card.get("scanned_by", ""),
                "Yes" if card.get("shared") else "No",
                uploaded_str,
            ]

            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num).value = value

        # Adjust column widths
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 25
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 30
        ws.column_dimensions["H"].width = 25
        ws.column_dimensions["I"].width = 20
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 20

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = "attachment; filename=all_cards.xlsx"
        response.headers["Content-Type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return response
    except Exception as e:
        print(f"Export Excel error: {e}")
        return jsonify({"error": f"Failed to export Excel: {str(e)}"}), 500


@app.route("/admin/export_pdf")
@login_required
@admin_required
def admin_export_pdf():
    """Export all cards as PDF (admin only)"""
    try:
        cards = list(collection.find({}))

        # Create PDF document
        output = io.BytesIO()
        pdf_doc = SimpleDocTemplate(
            output,
            pagesize=letter,
            topMargin=0.4 * inch,
            bottomMargin=0.4 * inch,
            leftMargin=0.4 * inch,
            rightMargin=0.4 * inch,
        )

        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=16,
            textColor=colors.HexColor("#667eea"),
            spaceAfter=4,
            alignment=1,  # Center
        )

        story = [
            Paragraph("Visiting Cards Report", title_style),
            Paragraph(f"Total Cards: {len(cards)}", styles["Normal"]),
            Spacer(1, 0.1 * inch),
        ]

        # Prepare table data - use text strings for simpler rendering
        headers = ["Name", "Company", "Email", "Phone", "Scanned By", "Shared"]
        table_data = [headers]

        for card in cards:
            row = [
                str(card.get("name", "N/A"))[:50],
                str(card.get("company", "N/A"))[:40],
                str(card.get("email", "N/A"))[:45],
                str(card.get("phone", "N/A"))[:20],
                str(card.get("scanned_by", "N/A"))[:30],
                "Yes" if card.get("shared") else "No",
            ]
            table_data.append(row)

        # Create table with proper column widths
        pdf_table = Table(
            table_data,
            colWidths=[
                1.0 * inch,
                1.2 * inch,
                1.5 * inch,
                1.0 * inch,
                1.3 * inch,
                0.6 * inch,
            ],
        )

        # Enhanced table styling with proper spacing and wrapping support
        table_style = TableStyle(
            [
                # Header styling
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#667eea")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 8),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, 0), 5),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
                # Data rows styling
                ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 1), (-1, -1), 6),
                ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                ("VALIGN", (0, 1), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 1), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
                ("LEFTPADDING", (0, 1), (-1, -1), 2),
                ("RIGHTPADDING", (0, 1), (-1, -1), 2),
                # Grid lines
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                (
                    "ROWBACKGROUNDS",
                    (0, 1),
                    (-1, -1),
                    [colors.white, colors.HexColor("#f5f5f5")],
                ),
                # Center align the "Shared" column
                ("ALIGN", (5, 1), (5, -1), "CENTER"),
                # Minimum height for rows
                ("MINHEIGHT", (0, 1), (-1, -1), 0.25 * inch),
            ]
        )

        pdf_table.setStyle(table_style)
        story.append(pdf_table)

        # Build PDF
        pdf_doc.build(story)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = "attachment; filename=all_cards.pdf"
        response.headers["Content-Type"] = "application/pdf"
        return response
    except Exception as e:
        print(f"Export PDF error: {e}")
        import traceback

        traceback.print_exc()
        return jsonify({"error": f"Failed to export PDF: {str(e)}"}), 500


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5000)
